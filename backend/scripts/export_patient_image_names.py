#!/usr/bin/env python3
"""Export image filenames for patients with an exact name match."""

import argparse
import sys
from collections.abc import Sequence
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
BACKEND_ROOT = SCRIPT_DIR.parent
sys.path.insert(0, str(BACKEND_ROOT))

from scripts.env_loader import load_project_env  # noqa: E402 - path bootstrap above.

load_project_env()

from openpyxl import Workbook  # noqa: E402 - environment must load before app imports.
from sqlalchemy import select  # noqa: E402
from sqlalchemy.exc import SQLAlchemyError  # noqa: E402
from sqlalchemy.orm import Session  # noqa: E402

from app.models.image_file import ImageFile  # noqa: E402
from app.models.patient import Patient  # noqa: E402
from app.shared.database import SessionLocal  # noqa: E402

DEFAULT_PATIENT_NAME = "中国青少年标准脊柱序列研究"
DEFAULT_OUTPUT_NAME = f"{DEFAULT_PATIENT_NAME}_影像名.xlsx"


def fetch_image_filenames(
    db: Session,
    patient_name: str,
    *,
    include_deleted: bool = False,
) -> list[str]:
    """Return original filenames for every patient matching the exact name."""
    statement = (
        select(ImageFile.original_filename)
        .join(Patient, ImageFile.patient_id == Patient.id)
        .where(Patient.name == patient_name)
        .order_by(ImageFile.id.asc())
    )

    if not include_deleted:
        statement = statement.where(
            Patient.is_deleted.is_(False),
            ImageFile.is_deleted.is_(False),
        )

    return list(db.execute(statement).scalars().all())


def write_filenames_excel(
    filenames: Sequence[str],
    output_path: Path,
) -> None:
    """Write filenames to a one-column XLSX workbook."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    if worksheet is None:
        raise RuntimeError("无法创建 Excel 工作表")
    worksheet.title = "影像名"
    worksheet.column_dimensions["A"].width = 48
    worksheet.cell(row=1, column=1, value="影像名")

    for row_index, filename in enumerate(filenames, start=2):
        cell = worksheet.cell(row=row_index, column=1, value=filename)
        # Force string storage so filenames beginning with "=" are not Excel formulas.
        cell.data_type = "s"

    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="按患者姓名精确匹配，并将关联影像原始文件名导出为单列 Excel。",
    )
    parser.add_argument(
        "--patient-name",
        default=DEFAULT_PATIENT_NAME,
        help=f"患者姓名，默认：{DEFAULT_PATIENT_NAME}",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(DEFAULT_OUTPUT_NAME),
        help=f"输出 XLSX 路径，默认：{DEFAULT_OUTPUT_NAME}",
    )
    parser.add_argument(
        "--include-deleted",
        action="store_true",
        help="同时包含已软删除的患者和影像。",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    try:
        with SessionLocal() as db:
            filenames = fetch_image_filenames(
                db,
                args.patient_name,
                include_deleted=args.include_deleted,
            )
    except SQLAlchemyError as error:
        print(
            f"数据库查询失败，请检查数据库连接配置：{error.__class__.__name__}",
            file=sys.stderr,
        )
        return 2

    if not filenames:
        print(f"未找到患者名称为“{args.patient_name}”的关联影像。")
        return 1

    output_path = args.output.expanduser().resolve()
    write_filenames_excel(filenames, output_path)
    print(f"已导出 {len(filenames)} 条影像名：{output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
