"""Excel adapter dedicated to the filename-driven dataset export use case."""

from __future__ import annotations

import os
import tempfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.contexts.imaging.application.dto import DatasetExportItemResult
from app.contexts.imaging.domain.exports import MEASUREMENT_COLUMN_ALIASES

_MANIFEST_SHEET = "export_manifest"
_MANIFEST_HEADERS = (
    "Excel行号",
    "请求文件名",
    "影像ID",
    "PatientID",
    "同名候选数",
    "候选影像ID",
    "检查类型",
    "输出路径",
    "关键点数量",
    "测量覆盖",
    "导出状态",
    "说明",
)


class DatasetExportWorkbook:
    """Preserve the caller workbook while filling results and an audit manifest."""

    def __init__(self, workbook: Workbook, worksheet: Worksheet) -> None:
        self._workbook = workbook
        self._worksheet = worksheet
        self._headers = self._header_columns(worksheet)
        if "影像名称" not in self._headers:
            raise ValueError("Sheet1 缺少“影像名称”列")
        missing_columns = [
            column
            for column in MEASUREMENT_COLUMN_ALIASES
            if column not in self._headers
        ]
        if missing_columns:
            raise ValueError(f"Sheet1 缺少测量列: {', '.join(missing_columns)}")

    @classmethod
    def load(cls, input_path: Path) -> DatasetExportWorkbook:
        workbook = load_workbook(input_path)
        if "Sheet1" not in workbook.sheetnames:
            raise ValueError("输入工作簿缺少 Sheet1")
        return cls(workbook, workbook["Sheet1"])

    def requested_rows(self) -> list[tuple[int, str]]:
        filename_column = self._headers["影像名称"]
        rows: list[tuple[int, str]] = []
        for row_number in range(2, self._worksheet.max_row + 1):
            value = self._worksheet.cell(
                row=row_number,
                column=filename_column,
            ).value
            if value is not None and str(value).strip():
                rows.append((row_number, str(value).strip()))
        return rows

    def clear_measurements(self, row_number: int) -> None:
        for column in MEASUREMENT_COLUMN_ALIASES:
            self._worksheet.cell(
                row=row_number,
                column=self._headers[column],
            ).value = None

    def write_measurements(self, row_number: int, values: dict[str, float]) -> None:
        for column, value in values.items():
            cell = self._worksheet.cell(
                row=row_number,
                column=self._headers[column],
            )
            cell.value = value
            cell.number_format = "0.00"

    def write_manifest(self, results: list[DatasetExportItemResult]) -> None:
        if _MANIFEST_SHEET in self._workbook.sheetnames:
            self._workbook.remove(self._workbook[_MANIFEST_SHEET])
        manifest = self._workbook.create_sheet(_MANIFEST_SHEET)
        manifest.append(_MANIFEST_HEADERS)
        for result in results:
            manifest.append(
                (
                    result.row_number,
                    result.requested_filename,
                    result.image_file_id,
                    result.patient_identifier,
                    result.candidate_count,
                    ",".join(str(item_id) for item_id in result.candidate_ids),
                    result.exam_type,
                    result.output_path,
                    result.keypoint_count,
                    f"{result.measurement_coverage}/{len(MEASUREMENT_COLUMN_ALIASES)}",
                    result.status,
                    result.detail,
                )
            )
        widths = (12, 48, 12, 20, 14, 30, 18, 64, 14, 14, 22, 72)
        for index, width in enumerate(widths, start=1):
            column = manifest.cell(row=1, column=index).column_letter
            manifest.column_dimensions[column].width = width

    def save_atomic(self, output_path: Path) -> None:
        temporary = tempfile.NamedTemporaryFile(
            prefix=f".{output_path.stem}-",
            suffix=".xlsx",
            dir=output_path.parent,
            delete=False,
        )
        temporary_path = Path(temporary.name)
        temporary.close()
        try:
            self._workbook.save(temporary_path)
            os.replace(temporary_path, output_path)
        finally:
            temporary_path.unlink(missing_ok=True)

    @staticmethod
    def _header_columns(worksheet: Worksheet) -> dict[str, int]:
        return {
            str(cell.value).strip(): int(cell.column)
            for cell in worksheet[1]
            if cell.value is not None
        }
