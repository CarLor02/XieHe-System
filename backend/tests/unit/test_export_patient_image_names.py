from pathlib import Path

from openpyxl import load_workbook

from scripts.export_patient_image_names import (
    fetch_image_filenames,
    write_filenames_excel,
)


class ScalarResult:
    def all(self) -> list[str]:
        return ["first.png", "same-name.png", "same-name.png"]


class ExecuteResult:
    def scalars(self) -> ScalarResult:
        return ScalarResult()


class RecordingSession:
    def __init__(self) -> None:
        self.sql = ""

    def execute(self, statement: object) -> ExecuteResult:
        self.sql = str(
            statement.compile(compile_kwargs={"literal_binds": True})
        )
        return ExecuteResult()


def test_fetch_image_filenames_uses_exact_name_and_active_rows() -> None:
    session = RecordingSession()

    filenames = fetch_image_filenames(
        session,  # type: ignore[arg-type]
        "中国青少年标准脊柱序列研究",
    )

    assert filenames == ["first.png", "same-name.png", "same-name.png"]
    assert "JOIN patients ON image_files.patient_id = patients.id" in session.sql
    assert "patients.name = '中国青少年标准脊柱序列研究'" in session.sql
    assert "patients.is_deleted IS false" in session.sql
    assert "image_files.is_deleted IS false" in session.sql
    assert "ORDER BY image_files.id ASC" in session.sql


def test_write_filenames_excel_creates_one_column_and_preserves_duplicates(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "image-names.xlsx"

    write_filenames_excel(
        ["first.png", "same-name.png", "same-name.png", "=image.png"],
        output_path,
    )

    workbook = load_workbook(output_path)
    worksheet = workbook["影像名"]

    assert worksheet.max_column == 1
    assert [cell.value for cell in worksheet["A"]] == [
        "影像名",
        "first.png",
        "same-name.png",
        "same-name.png",
        "=image.png",
    ]
    assert worksheet["A5"].data_type == "s"
