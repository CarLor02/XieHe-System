from __future__ import annotations

import io
import json
from pathlib import Path
from typing import Any, BinaryIO, cast

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image

from app.contexts.imaging.application.dto import DatasetExportCandidate, StoredObject
from app.contexts.imaging.application.errors import (
    DatasetExportTeamAmbiguousError,
    DatasetExportTeamNotFoundError,
    ObjectStorageObjectNotFoundError,
)
from app.contexts.imaging.application.exports import DatasetExportService
from app.contexts.imaging.domain import JsonObject
from app.contexts.imaging.domain.exports import MEASUREMENT_COLUMN_ALIASES


def _image_bytes(image_format: str, *, size: tuple[int, int]) -> bytes:
    output = io.BytesIO()
    Image.new("L", size, color=127).save(output, format=image_format)
    return output.getvalue()


def _candidate(
    image_file_id: int,
    filename: str,
    *,
    patient_identifier: str,
    object_key: str,
    file_size: int,
    annotation: JsonObject | None,
) -> DatasetExportCandidate:
    return DatasetExportCandidate(
        image_file_id=image_file_id,
        original_filename=filename,
        description="侧位X光片",
        storage_bucket="medical-images",
        object_key=object_key,
        file_size=file_size,
        patient_identifier=patient_identifier,
        annotation=annotation,
    )


class FakeDatasetExportRepository:
    def __init__(
        self,
        candidates: list[DatasetExportCandidate],
        *,
        team_ids_by_name: dict[str, list[int]] | None = None,
    ) -> None:
        self.candidates = candidates
        self.team_ids_by_name = team_ids_by_name or {}
        self.resolved_team_names: list[str] = []
        self.requested_team_ids: list[int | None] = []

    def find_active_team_ids_by_exact_name(self, team_name: str) -> list[int]:
        self.resolved_team_names.append(team_name)
        return self.team_ids_by_name.get(team_name, [])

    def find_candidates(
        self,
        *,
        filenames: list[str],
        exam_type: str,
        team_id: int | None,
    ) -> list[DatasetExportCandidate]:
        assert exam_type == "侧位X光片"
        self.requested_team_ids.append(team_id)
        return [item for item in self.candidates if item.original_filename in filenames]


class FakeObjectStorage:
    def __init__(self, payloads: dict[str, bytes]) -> None:
        self.payloads = payloads

    async def stat_object(self, *, bucket: str, object_key: str) -> StoredObject:
        assert bucket == "medical-images"
        try:
            payload = self.payloads[object_key]
        except KeyError as exc:
            raise ObjectStorageObjectNotFoundError(object_key) from exc
        return StoredObject(size=len(payload), etag=f"etag-{object_key}", metadata={})

    async def download_object_to(
        self,
        *,
        bucket: str,
        object_key: str,
        destination: BinaryIO,
    ) -> None:
        assert bucket == "medical-images"
        destination.write(self.payloads[object_key])


class DisappearingObjectStorage(FakeObjectStorage):
    async def download_object_to(
        self,
        *,
        bucket: str,
        object_key: str,
        destination: BinaryIO,
    ) -> None:
        raise ObjectStorageObjectNotFoundError(
            f"{bucket}/{object_key} disappeared after stat"
        )


def _create_input_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = cast(Worksheet, workbook.active)
    worksheet.title = "Sheet1"
    headers = [
        "No.",
        "影像名称",
        *MEASUREMENT_COLUMN_ALIASES,
        "备注",
    ]
    worksheet.append(headers)
    worksheet.append([1, "same.jpg", *([None] * 12), "保留备注 A"])
    worksheet.append([2, "empty.jpg", *([None] * 12), "保留备注 B"])
    worksheet.append([3, "absent.png", *([None] * 12), "保留备注 C"])
    worksheet["A1"].fill = PatternFill(fill_type="solid", fgColor="00FF00")
    worksheet.column_dimensions["B"].width = 42
    workbook.save(path)


@pytest.mark.asyncio
async def test_dataset_export_continues_failures_and_preserves_workbook(
    tmp_path: Path,
) -> None:
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output"
    _create_input_workbook(input_path)
    png_payload = _image_bytes("PNG", size=(100, 200))
    jpeg_payload = _image_bytes("JPEG", size=(80, 120))
    annotation: JsonObject = {
        "imageWidth": 100,
        "imageHeight": 200,
        "vertebraeLayer": [
            {
                "label": "T1",
                "corners": [
                    {"x": 10, "y": 20},
                    {"x": 30, "y": 20},
                    {"x": 10, "y": 60},
                    {"x": 30, "y": 60},
                ],
            }
        ],
        "measurements": [
            {"type": "T1 Slope", "value": "12.00°"},
            {"type": "t1-slope", "value": "42.35°"},
            {"type": "PT", "value": "-5.50°"},
        ],
    }
    candidates = [
        _candidate(
            20,
            "same.jpg",
            patient_identifier="P001",
            object_key="missing.jpg",
            file_size=999,
            annotation=annotation,
        ),
        _candidate(
            10,
            "same.jpg",
            patient_identifier="P001",
            object_key="valid.png",
            file_size=len(png_payload),
            annotation=annotation,
        ),
        _candidate(
            30,
            "empty.jpg",
            patient_identifier="P002",
            object_key="empty.jpg",
            file_size=len(jpeg_payload),
            annotation={"measurements": []},
        ),
    ]
    repository = FakeDatasetExportRepository(
        candidates,
        team_ids_by_name={"脊柱研究团队": [7]},
    )
    service = DatasetExportService(
        cast(Any, repository),
        cast(
            Any,
            FakeObjectStorage({"valid.png": png_payload, "empty.jpg": jpeg_payload}),
        ),
    )

    result = await service.run(
        input_workbook=input_path,
        output_directory=output_path,
        exam_type="侧位X光片",
        overwrite=False,
        team_name=" 脊柱研究团队 ",
    )

    assert repository.resolved_team_names == ["脊柱研究团队"]
    assert repository.requested_team_ids == [7]

    assert result.summary.requested == 3
    assert result.summary.succeeded == 2
    assert result.summary.empty_annotations == 1
    assert result.summary.not_found == 1
    assert result.summary.duplicate_overwrites == 1
    assert result.summary.failed == 1
    assert result.has_failures is True

    exported_png = output_path / "P001/same.png"
    assert exported_png.read_bytes() == png_payload
    labelme = json.loads((output_path / "P001/same.json").read_text("utf-8"))
    assert labelme["imagePath"] == "same.png"
    assert labelme["shapes"][0]["shape_type"] == "polygon"
    assert labelme["shapes"][0]["points"] == [
        [10.0, 20.0],
        [30.0, 20.0],
        [30.0, 60.0],
        [10.0, 60.0],
    ]

    with Image.open(output_path / "P002/empty.png") as converted:
        assert converted.format == "PNG"
        assert converted.size == (80, 120)
    empty_labelme = json.loads((output_path / "P002/empty.json").read_text("utf-8"))
    assert empty_labelme["shapes"] == []

    measurements = load_workbook(output_path / "measurements.xlsx")
    sheet = measurements["Sheet1"]
    headers = {
        str(cell.value): int(cell.column) for cell in sheet[1] if cell.value is not None
    }
    assert sheet.cell(2, headers["T1 slope"]).value == 42.35
    assert sheet.cell(2, headers["PT"]).value == -5.5
    assert sheet.cell(2, headers["T12-L1"]).value is None
    assert sheet.cell(2, headers["备注"]).value == "保留备注 A"
    assert sheet["A1"].fill.fgColor.rgb == "0000FF00"
    assert sheet.column_dimensions["B"].width == 42

    manifest = measurements["export_manifest"]
    manifest_headers = {
        str(cell.value): int(cell.column)
        for cell in manifest[1]
        if cell.value is not None
    }
    assert manifest.cell(2, manifest_headers["影像ID"]).value == 10
    assert manifest.cell(2, manifest_headers["候选影像ID"]).value == "20,10"
    assert manifest.cell(2, manifest_headers["关键点数量"]).value == 4
    assert manifest.cell(3, manifest_headers["导出状态"]).value == "missing_annotation"
    assert manifest.cell(4, manifest_headers["导出状态"]).value == "not_found"

    summary = json.loads((output_path / "export-summary.json").read_text("utf-8"))
    assert summary == {
        "requested": 3,
        "succeeded": 2,
        "empty_annotations": 1,
        "not_found": 1,
        "duplicate_overwrites": 1,
        "object_missing": 0,
        "failed": 1,
    }


@pytest.mark.asyncio
async def test_dataset_export_classifies_object_removed_after_stat(
    tmp_path: Path,
) -> None:
    input_path = tmp_path / "input.xlsx"
    _create_input_workbook(input_path)
    workbook = load_workbook(input_path)
    sheet = workbook["Sheet1"]
    sheet.delete_rows(3, 2)
    workbook.save(input_path)
    payload = _image_bytes("PNG", size=(16, 32))
    candidate = _candidate(
        10,
        "same.jpg",
        patient_identifier="P001",
        object_key="disappearing.png",
        file_size=len(payload),
        annotation={"vertebraeLayer": []},
    )
    service = DatasetExportService(
        cast(Any, FakeDatasetExportRepository([candidate])),
        cast(Any, DisappearingObjectStorage({"disappearing.png": payload})),
    )

    result = await service.run(
        input_workbook=input_path,
        output_directory=tmp_path / "output",
        exam_type="侧位X光片",
        overwrite=False,
    )

    assert result.summary.object_missing == 1
    assert result.summary.failed == 1
    assert result.items[0].status == "object_missing"
    assert not (tmp_path / "output/P001/same.png").exists()


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("team_ids", "expected_error"),
    [
        ([], DatasetExportTeamNotFoundError),
        ([1, 2], DatasetExportTeamAmbiguousError),
    ],
)
async def test_dataset_export_rejects_unresolved_team_name(
    tmp_path: Path,
    team_ids: list[int],
    expected_error: type[Exception],
) -> None:
    input_path = tmp_path / "input.xlsx"
    _create_input_workbook(input_path)
    repository = FakeDatasetExportRepository(
        [],
        team_ids_by_name={"重名团队": team_ids},
    )
    service = DatasetExportService(
        cast(Any, repository),
        cast(Any, FakeObjectStorage({})),
    )

    with pytest.raises(expected_error):
        await service.run(
            input_workbook=input_path,
            output_directory=tmp_path / "output",
            exam_type="侧位X光片",
            overwrite=False,
            team_name="重名团队",
        )

    assert repository.resolved_team_names == ["重名团队"]
    assert repository.requested_team_ids == []
